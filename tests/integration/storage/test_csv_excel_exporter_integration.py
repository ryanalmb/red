"""Integration tests for CSV/Excel Exporter (Story 13.8).

Tests real CSV and Excel export with actual pandas and openpyxl libraries.
No mocks - tests actual production code behavior.

Follows TDD: Write failing tests first (RED), then implement (GREEN).
"""

from __future__ import annotations

import csv
import io
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pytest

# Import dependencies - these should exist
from cyberred.storage.report_generator import ReportData, TimelineEvent

# Import the module under test - this will fail initially (RED phase)
from cyberred.storage.csv_excel_exporter import (
    CSVExporter,
    ExcelExporter,
    export_findings_csv,
    export_findings_xlsx,
)


# =============================================================================
# Test Fixtures - Realistic Engagement Data
# =============================================================================


@pytest.fixture
def realistic_findings() -> tuple[dict[str, Any], ...]:
    """Realistic findings from a penetration test engagement."""
    return (
        {
            "id": "finding-001",
            "type": "sqli",
            "severity": "critical",
            "target": "https://webapp.example.com/api/users?id=1",
            "evidence": """SQL Injection vulnerability discovered in user lookup API.
            
Payload: ' OR '1'='1' --
Response indicates successful injection with database error disclosure.

Impact: Attacker can bypass authentication and extract sensitive data.""",
            "timestamp": "2026-02-12T06:15:32.123Z",
            "agent_id": "webapp-agent-1",
            "tool": "sqlmap",
            "topic": "webapp",
            "attck_id": "T1190",
        },
        {
            "id": "finding-002",
            "type": "xss",
            "severity": "high",
            "target": "https://webapp.example.com/search",
            "evidence": """Reflected Cross-Site Scripting (XSS) in search parameter.

Payload: <script>alert('XSS')</script>
The payload is reflected without encoding in the HTML response.

Impact: Attacker can execute arbitrary JavaScript in victim's browser.""",
            "timestamp": "2026-02-12T06:45:18.456Z",
            "agent_id": "webapp-agent-1",
            "tool": "nuclei",
            "topic": "webapp",
            "attck_id": "T1059.007",
        },
        {
            "id": "finding-003",
            "type": "default_creds",
            "severity": "critical",
            "target": "192.168.1.50:22",
            "evidence": """SSH access with default credentials.

Username: admin
Password: admin123

Full shell access obtained to production server.""",
            "timestamp": "2026-02-12T07:22:45.789Z",
            "agent_id": "credential-agent-1",
            "tool": "hydra",
            "topic": "network",
            "attck_id": "T1078",
        },
        {
            "id": "finding-004",
            "type": "info_disclosure",
            "severity": "medium",
            "target": "https://webapp.example.com/",
            "evidence": "Server version disclosed in HTTP headers: Apache/2.4.41 (Ubuntu)",
            "timestamp": "2026-02-12T06:05:12.000Z",
            "agent_id": "recon-agent-1",
            "tool": "nmap",
            "topic": "network",
        },
        {
            "id": "finding-005",
            "type": "ssl_weak",
            "severity": "low",
            "target": "webapp.example.com:443",
            "evidence": "TLS 1.0 and TLS 1.1 are enabled. Recommend disabling in favor of TLS 1.2+",
            "timestamp": "2026-02-12T06:08:33.000Z",
            "agent_id": "recon-agent-1",
            "tool": "nmap",
            "topic": "network",
        },
        {
            "id": "finding-006",
            "type": "open_port",
            "severity": "info",
            "target": "192.168.1.1",
            "evidence": "Port 80 (HTTP) open - web server detected",
            "timestamp": "2026-02-12T06:02:00.000Z",
            "agent_id": "recon-agent-1",
            "tool": "nmap",
            "topic": "network",
        },
    )


@pytest.fixture
def realistic_report_data(realistic_findings: tuple[dict[str, Any], ...]) -> ReportData:
    """Realistic ReportData from a penetration test engagement."""
    return ReportData(
        engagement_id="eng-2026-0212-001",
        title="External Penetration Test - Example Corp",
        start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
        end_time=datetime(2026, 2, 12, 18, 0, 0, tzinfo=timezone.utc),
        scope={
            "targets": ["192.168.1.0/24", "webapp.example.com", "api.example.com"],
            "exclusions": ["192.168.1.254"],
        },
        findings=realistic_findings,
        timeline_events=(
            TimelineEvent(
                timestamp="2026-02-12T06:00:00Z",
                event_type="engagement_start",
                description="Engagement started",
                agent_id="system",
            ),
            TimelineEvent(
                timestamp="2026-02-12T06:15:32Z",
                event_type="finding_discovered",
                description="SQL Injection discovered",
                agent_id="webapp-agent-1",
            ),
            TimelineEvent(
                timestamp="2026-02-12T18:00:00Z",
                event_type="engagement_end",
                description="Engagement completed",
                agent_id="system",
            ),
        ),
    )


# =============================================================================
# Integration Test: Full CSV Cycle
# =============================================================================


class TestCSVIntegration:
    """Integration tests for CSV export - no mocks, real behavior."""

    def test_full_csv_cycle_export_and_parse(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test full CSV cycle: create ReportData with findings → export → parse back."""
        exporter = CSVExporter()
        
        # Export to CSV string
        csv_content = exporter.export(realistic_report_data)
        
        # Parse back with Python csv module
        reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(reader)
        
        # Verify all findings are present
        assert len(rows) == 6
        
        # Verify first finding data
        first_row = rows[0]
        assert first_row["severity"] == "critical"
        assert first_row["type"] == "sqli"
        assert "webapp.example.com" in first_row["target"]
        assert "SQL Injection" in first_row["description"]

    def test_csv_round_trip_preserves_data(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test CSV round-trip preserves data accurately."""
        exporter = CSVExporter()
        
        # Export
        csv_content = exporter.export(realistic_report_data)
        
        # Parse back
        reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(reader)
        
        # Verify each finding's data matches original
        for i, row in enumerate(rows):
            original = realistic_report_data.findings[i]
            assert row["severity"] == original["severity"]
            assert row["type"] == original["type"]
            assert row["target"] == original["target"]
            # evidence maps to description column
            assert row["description"] == original["evidence"]

    def test_csv_file_io_real_filesystem(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test file I/O with real filesystem."""
        exporter = CSVExporter()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "findings.csv"
            
            # Export to file
            exporter.export(realistic_report_data, output_path=output_path)
            
            # Verify file exists
            assert output_path.exists()
            
            # Read back and verify
            content = output_path.read_text(encoding="utf-8")
            assert "severity,type,target,description,timestamp" in content
            assert "critical" in content
            assert "sqli" in content

    def test_csv_multiline_evidence_preserved(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test multiline evidence text is preserved in CSV."""
        exporter = CSVExporter()
        
        csv_content = exporter.export(realistic_report_data)
        
        # Parse back
        reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(reader)
        
        # Find finding with multiline evidence (SQL injection)
        sqli_finding = next(r for r in rows if r["type"] == "sqli")
        
        # Verify newlines are preserved
        assert "\n" in sqli_finding["description"]
        assert "Payload:" in sqli_finding["description"]
        assert "Impact:" in sqli_finding["description"]

    def test_csv_extended_columns_integration(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test extended columns with real data."""
        exporter = CSVExporter()
        
        # Export with extended columns
        csv_content = exporter.export(realistic_report_data, extended=True)
        
        # Parse back
        reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(reader)
        
        # Verify extended columns present
        first_row = rows[0]
        assert "agent_id" in first_row
        assert "tool" in first_row
        assert "topic" in first_row
        assert "attck_id" in first_row
        
        # Verify values
        assert first_row["agent_id"] == "webapp-agent-1"
        assert first_row["tool"] == "sqlmap"


# =============================================================================
# Integration Test: Full Excel Cycle
# =============================================================================


class TestExcelIntegration:
    """Integration tests for Excel export - no mocks, real behavior."""

    def test_full_excel_cycle_export_and_read(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test full Excel cycle: create ReportData with findings → export → read with openpyxl."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        
        # Export to Excel bytes
        xlsx_bytes = exporter.export(realistic_report_data)
        
        # Read back with openpyxl
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        
        # Verify all findings are present (header + 6 findings)
        assert ws.max_row == 7
        
        # Verify header
        headers = [cell.value for cell in ws[1]]
        assert headers == ["severity", "type", "target", "description", "timestamp"]
        
        # Verify first finding data
        assert ws.cell(row=2, column=1).value == "critical"
        assert ws.cell(row=2, column=2).value == "sqli"

    def test_excel_round_trip_preserves_data(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test Excel round-trip preserves data accurately."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        
        # Export
        xlsx_bytes = exporter.export(realistic_report_data)
        
        # Read back
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        
        # Verify each finding's data matches original
        for i, finding in enumerate(realistic_report_data.findings):
            row_num = i + 2  # Skip header
            assert ws.cell(row=row_num, column=1).value == finding["severity"]
            assert ws.cell(row=row_num, column=2).value == finding["type"]
            assert ws.cell(row=row_num, column=3).value == finding["target"]
            # evidence maps to description column
            assert ws.cell(row=row_num, column=4).value == finding["evidence"]

    def test_excel_file_io_real_filesystem(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test file I/O with real filesystem."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "findings.xlsx"
            
            # Export to file
            exporter.export(realistic_report_data, output_path=output_path)
            
            # Verify file exists
            assert output_path.exists()
            
            # Read back with openpyxl
            workbook = load_workbook(output_path)
            ws = workbook.active
            
            assert ws.max_row == 7  # Header + 6 findings

    def test_excel_formatting_preserved(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test Excel formatting is applied correctly."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        xlsx_bytes = exporter.export(realistic_report_data)
        
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        
        # Verify worksheet name
        assert ws.title == "Findings"
        
        # Verify header is bold
        for cell in ws[1]:
            assert cell.font.bold is True
        
        # Verify auto-filter
        assert ws.auto_filter.ref is not None

    def test_excel_extended_columns_integration(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test extended columns with real data in Excel."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        
        # Export with extended columns
        xlsx_bytes = exporter.export(realistic_report_data, extended=True)
        
        # Read back
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        
        # Verify extended headers
        headers = [cell.value for cell in ws[1]]
        expected = [
            "severity", "type", "target", "description", "timestamp",
            "agent_id", "tool", "topic", "attck_id"
        ]
        assert headers == expected
        
        # Verify extended values
        assert ws.cell(row=2, column=6).value == "webapp-agent-1"
        assert ws.cell(row=2, column=7).value == "sqlmap"
        assert ws.cell(row=2, column=8).value == "webapp"
        assert ws.cell(row=2, column=9).value == "T1190"


# =============================================================================
# Integration Test: Convenience Functions
# =============================================================================


class TestConvenienceFunctionsIntegration:
    """Integration tests for convenience functions."""

    def test_export_findings_csv_integration(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test export_findings_csv with realistic data."""
        csv_content = export_findings_csv(realistic_report_data)
        
        # Parse and verify
        reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(reader)
        
        assert len(rows) == 6
        assert rows[0]["severity"] == "critical"

    def test_export_findings_xlsx_integration(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test export_findings_xlsx with realistic data."""
        from openpyxl import load_workbook
        
        xlsx_bytes = export_findings_xlsx(realistic_report_data)
        
        # Load and verify
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        
        assert ws.max_row == 7  # Header + 6 findings

    def test_csv_and_xlsx_data_consistency(
        self, realistic_report_data: ReportData
    ) -> None:
        """Test CSV and Excel export produce consistent data."""
        from openpyxl import load_workbook
        
        # Export both formats
        csv_content = export_findings_csv(realistic_report_data)
        xlsx_bytes = export_findings_xlsx(realistic_report_data)
        
        # Parse CSV
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        csv_rows = list(csv_reader)
        
        # Parse Excel
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        
        # Compare data
        for i, csv_row in enumerate(csv_rows):
            excel_row = i + 2  # Skip header
            assert csv_row["severity"] == ws.cell(row=excel_row, column=1).value
            assert csv_row["type"] == ws.cell(row=excel_row, column=2).value
            assert csv_row["target"] == ws.cell(row=excel_row, column=3).value
            assert csv_row["description"] == ws.cell(row=excel_row, column=4).value


# =============================================================================
# Integration Test: Edge Cases with Real Libraries
# =============================================================================


class TestEdgeCasesIntegration:
    """Integration tests for edge cases with real libraries."""

    def test_unicode_handling_csv(self) -> None:
        """Test Unicode characters in CSV with real encoding."""
        findings = (
            {
                "severity": "high",
                "type": "xss",
                "target": "http://example.com/search?q=日本語",
                "evidence": "XSS with Unicode: <script>alert('中文 🔥 한국어')</script>",
                "timestamp": "2026-02-12T06:00:00Z",
            },
        )
        report_data = ReportData(
            engagement_id="eng-unicode-001",
            title="Unicode Test",
            start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
            end_time=None,
            scope={"targets": [], "exclusions": []},
            findings=findings,
            timeline_events=(),
        )
        
        exporter = CSVExporter()
        csv_content = exporter.export(report_data)
        
        # Verify Unicode preserved
        assert "日本語" in csv_content
        assert "中文" in csv_content
        assert "🔥" in csv_content
        assert "한국어" in csv_content
        
        # Verify it can be parsed back
        reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(reader)
        assert "日本語" in rows[0]["target"]

    def test_unicode_handling_excel(self) -> None:
        """Test Unicode characters in Excel with real libraries."""
        from openpyxl import load_workbook
        
        findings = (
            {
                "severity": "high",
                "type": "xss",
                "target": "http://example.com/search?q=日本語",
                "evidence": "XSS with Unicode: 中文 🔥 한국어",
                "timestamp": "2026-02-12T06:00:00Z",
            },
        )
        report_data = ReportData(
            engagement_id="eng-unicode-001",
            title="Unicode Test",
            start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
            end_time=None,
            scope={"targets": [], "exclusions": []},
            findings=findings,
            timeline_events=(),
        )
        
        exporter = ExcelExporter()
        xlsx_bytes = exporter.export(report_data)
        
        # Verify it can be read back with Unicode intact
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        
        target = ws.cell(row=2, column=3).value
        evidence = ws.cell(row=2, column=4).value
        
        assert "日本語" in target
        assert "中文" in evidence
        assert "한국어" in evidence

    def test_large_dataset_performance(self) -> None:
        """Test with 1000+ findings for performance."""
        from openpyxl import load_workbook
        
        findings = tuple(
            {
                "id": f"finding-{i}",
                "severity": ["critical", "high", "medium", "low", "info"][i % 5],
                "type": ["sqli", "xss", "rce", "ssrf", "idor"][i % 5],
                "target": f"http://example.com/endpoint/{i}",
                "evidence": f"Finding {i} with detailed evidence about the vulnerability discovered during testing.",
                "timestamp": f"2026-02-12T{(6 + i // 60) % 24:02d}:{i % 60:02d}:00Z",
                "agent_id": f"agent-{i % 10}",
                "tool": ["nmap", "nuclei", "sqlmap", "ffuf", "hydra"][i % 5],
                "topic": ["webapp", "network", "api"][i % 3],
            }
            for i in range(1000)
        )
        report_data = ReportData(
            engagement_id="eng-perf-001",
            title="Performance Test",
            start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
            end_time=None,
            scope={"targets": [], "exclusions": []},
            findings=findings,
            timeline_events=(),
        )
        
        # CSV export
        csv_exporter = CSVExporter()
        csv_content = csv_exporter.export(report_data)
        
        csv_reader = csv.reader(io.StringIO(csv_content))
        csv_rows = list(csv_reader)
        assert len(csv_rows) == 1001  # Header + 1000 findings
        
        # Excel export
        excel_exporter = ExcelExporter()
        xlsx_bytes = excel_exporter.export(report_data)
        
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        ws = workbook.active
        assert ws.max_row == 1001  # Header + 1000 findings
