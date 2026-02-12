"""Unit tests for CSV/Excel Exporter (Story 13.8).

Tests CSV and Excel export for spreadsheet analysis.
Follows TDD: Write failing tests first (RED), then implement (GREEN).

Acceptance Criteria:
1. Given engagement has findings
2. When I export with format=csv or format=xlsx
3. Then one row per finding with columns: severity, type, target, description, timestamp
4. And CSV uses UTF-8 encoding with proper escaping
5. And Excel includes formatted headers and auto-filter
6. And unit tests verify export accuracy
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pytest

# Import dependencies - these should exist
from cyberred.storage.report_generator import ReportData, TimelineEvent

# Import the module under test - this will fail initially (RED phase)
from cyberred.storage.csv_excel_exporter import (
    CSVExporter,
    ExcelExporter,
    export_findings_csv,
    export_findings_xlsx,
)


# =============================================================================
# Test Fixtures
# =============================================================================


@pytest.fixture
def sample_findings() -> tuple[dict[str, Any], ...]:
    """Sample findings with various severities and types."""
    return (
        {
            "id": "finding-001",
            "type": "sqli",
            "severity": "critical",
            "target": "http://192.168.1.100/login",
            "evidence": "SQL Injection detected in login form",
            "timestamp": "2026-02-12T06:00:00Z",
            "agent_id": "recon-agent-1",
            "tool": "sqlmap",
            "topic": "webapp",
            "attck_id": "T1190",
        },
        {
            "id": "finding-002",
            "type": "xss",
            "severity": "high",
            "target": "https://example.com/search",
            "evidence": "Reflected XSS in search parameter",
            "timestamp": "2026-02-12T06:30:00Z",
            "agent_id": "webapp-agent-1",
            "tool": "nuclei",
            "topic": "webapp",
            "attck_id": "T1059.007",
        },
        {
            "id": "finding-003",
            "type": "rce",
            "severity": "critical",
            "target": "192.168.1.50",
            "evidence": "Remote code execution via command injection",
            "timestamp": "2026-02-12T07:00:00Z",
            "agent_id": "exploit-agent-1",
            "tool": "metasploit",
            "topic": "network",
            "attck_id": "T1059",
        },
        {
            "id": "finding-004",
            "type": "info_disclosure",
            "severity": "medium",
            "target": "api.example.com",
            "evidence": "Server version disclosed in headers",
            "timestamp": "2026-02-12T07:30:00Z",
            "agent_id": "recon-agent-1",
            "tool": "nmap",
            "topic": "network",
        },
        {
            "id": "finding-005",
            "type": "open_port",
            "severity": "info",
            "target": "10.0.0.1",
            "evidence": "Port 22 (SSH) open",
            "timestamp": "2026-02-12T08:00:00Z",
            "agent_id": "recon-agent-1",
            "tool": "nmap",
            "topic": "network",
        },
    )


@pytest.fixture
def sample_report_data(sample_findings: tuple[dict[str, Any], ...]) -> ReportData:
    """Sample ReportData for testing."""
    return ReportData(
        engagement_id="eng-test-001",
        title="Penetration Test - CSV/Excel Export Test",
        start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
        end_time=datetime(2026, 2, 12, 12, 0, 0, tzinfo=timezone.utc),
        scope={"targets": ["192.168.1.0/24", "example.com"], "exclusions": []},
        findings=sample_findings,
        timeline_events=(
            TimelineEvent(
                timestamp="2026-02-12T06:00:00Z",
                event_type="engagement_start",
                description="Engagement started",
                agent_id="system",
            ),
        ),
    )


@pytest.fixture
def empty_report_data() -> ReportData:
    """ReportData with no findings."""
    return ReportData(
        engagement_id="eng-empty-001",
        title="Empty Engagement",
        start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
        end_time=None,
        scope={"targets": [], "exclusions": []},
        findings=(),
        timeline_events=(),
    )


@pytest.fixture
def edge_case_findings() -> tuple[dict[str, Any], ...]:
    """Findings with edge cases: None values, Unicode, newlines, commas."""
    return (
        {
            "id": "edge-001",
            "type": "sqli",
            "severity": "critical",
            "target": "http://example.com/path,with,commas",
            "evidence": 'Value with "quotes" inside',
            "timestamp": "2026-02-12T06:00:00Z",
        },
        {
            "id": "edge-002",
            "type": None,
            "severity": None,
            "target": None,
            "evidence": None,
            "timestamp": None,
        },
        {
            "id": "edge-003",
            "type": "xss",
            "severity": "high",
            "target": "http://example.com/unicode",
            "evidence": "Unicode test: émojis 🔥🎯, CJK 中文日本語한국어, symbols ™©®",
            "timestamp": "2026-02-12T07:00:00Z",
        },
        {
            "id": "edge-004",
            "type": "info",
            "severity": "low",
            "target": "http://example.com/newlines",
            "evidence": "Line 1\nLine 2\nLine 3",
            "timestamp": "2026-02-12T08:00:00Z",
        },
    )


@pytest.fixture
def edge_case_report_data(edge_case_findings: tuple[dict[str, Any], ...]) -> ReportData:
    """ReportData with edge case findings."""
    return ReportData(
        engagement_id="eng-edge-001",
        title="Edge Case Test",
        start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
        end_time=None,
        scope={"targets": [], "exclusions": []},
        findings=edge_case_findings,
        timeline_events=(),
    )


@pytest.fixture
def datetime_findings() -> tuple[dict[str, Any], ...]:
    """Findings with datetime objects instead of strings."""
    return (
        {
            "id": "dt-001",
            "type": "sqli",
            "severity": "critical",
            "target": "http://example.com",
            "evidence": "Test finding",
            "timestamp": datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
        },
    )


@pytest.fixture
def datetime_report_data(datetime_findings: tuple[dict[str, Any], ...]) -> ReportData:
    """ReportData with datetime timestamp objects."""
    return ReportData(
        engagement_id="eng-dt-001",
        title="Datetime Test",
        start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
        end_time=None,
        scope={"targets": [], "exclusions": []},
        findings=datetime_findings,
        timeline_events=(),
    )


# =============================================================================
# Task 1: Test File Structure (AC: #6)
# =============================================================================


class TestCSVExcelExporterImports:
    """Verify imports work correctly."""

    def test_csv_exporter_importable(self) -> None:
        """Test CSVExporter can be imported."""
        assert CSVExporter is not None

    def test_excel_exporter_importable(self) -> None:
        """Test ExcelExporter can be imported."""
        assert ExcelExporter is not None

    def test_convenience_functions_importable(self) -> None:
        """Test convenience functions can be imported."""
        assert export_findings_csv is not None
        assert export_findings_xlsx is not None


# =============================================================================
# Task 2: CSVExporter Class Tests (AC: #1, #2, #4)
# =============================================================================


class TestCSVExporterInit:
    """Tests for CSVExporter initialization."""

    def test_init_creates_instance(self) -> None:
        """Test CSVExporter.__init__() initializes correctly."""
        exporter = CSVExporter()
        assert exporter is not None
        assert isinstance(exporter, CSVExporter)


class TestCSVExporterExport:
    """Tests for CSVExporter.export() method."""

    def test_export_returns_string(
        self, sample_report_data: ReportData
    ) -> None:
        """Test export(report_data) returns CSV string."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        assert isinstance(result, str)
        assert len(result) > 0

    def test_export_to_file(
        self, sample_report_data: ReportData, tmp_path: Path
    ) -> None:
        """Test export(report_data, output_path) writes to file."""
        exporter = CSVExporter()
        output_path = tmp_path / "findings.csv"
        
        result = exporter.export(sample_report_data, output_path=output_path)
        
        assert output_path.exists()
        assert isinstance(result, str)
        content = output_path.read_text(encoding="utf-8")
        assert content == result

    def test_export_has_header_row(
        self, sample_report_data: ReportData
    ) -> None:
        """Test CSV has header row with correct columns."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        reader = csv.reader(io.StringIO(result))
        header = next(reader)
        
        assert header == ["severity", "type", "target", "description", "timestamp"]

    def test_export_uses_utf8_encoding(
        self, sample_report_data: ReportData, tmp_path: Path
    ) -> None:
        """Test CSV uses UTF-8 encoding."""
        exporter = CSVExporter()
        output_path = tmp_path / "findings.csv"
        
        exporter.export(sample_report_data, output_path=output_path)
        
        # Read as bytes and verify UTF-8
        content_bytes = output_path.read_bytes()
        content_str = content_bytes.decode("utf-8")
        assert isinstance(content_str, str)


# =============================================================================
# Task 3: CSV Column Mapping Tests (AC: #3)
# =============================================================================


class TestCSVColumnMapping:
    """Tests for CSV column mapping."""

    def test_header_row_columns(
        self, sample_report_data: ReportData
    ) -> None:
        """Test first row is header: severity,type,target,description,timestamp."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        lines = result.strip().split("\n")
        assert lines[0] == "severity,type,target,description,timestamp"

    def test_each_finding_produces_one_row(
        self, sample_report_data: ReportData
    ) -> None:
        """Test each finding produces one row."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Header + 5 findings = 6 rows
        assert len(rows) == 6

    def test_severity_column_values(
        self, sample_report_data: ReportData
    ) -> None:
        """Test severity column contains finding severity."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Check severity values (column 0)
        severities = [row[0] for row in rows[1:]]  # Skip header
        assert "critical" in severities
        assert "high" in severities
        assert "medium" in severities
        assert "info" in severities

    def test_type_column_values(
        self, sample_report_data: ReportData
    ) -> None:
        """Test type column contains finding type."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Check type values (column 1)
        types = [row[1] for row in rows[1:]]
        assert "sqli" in types
        assert "xss" in types
        assert "rce" in types

    def test_target_column_values(
        self, sample_report_data: ReportData
    ) -> None:
        """Test target column contains target URL/IP."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Check target values (column 2)
        targets = [row[2] for row in rows[1:]]
        assert "http://192.168.1.100/login" in targets
        assert "https://example.com/search" in targets

    def test_description_column_values(
        self, sample_report_data: ReportData
    ) -> None:
        """Test description column contains evidence text."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Check description values (column 3)
        descriptions = [row[3] for row in rows[1:]]
        assert any("SQL Injection" in d for d in descriptions)

    def test_timestamp_column_values(
        self, sample_report_data: ReportData
    ) -> None:
        """Test timestamp column contains ISO timestamp."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Check timestamp values (column 4)
        timestamps = [row[4] for row in rows[1:]]
        assert "2026-02-12T06:00:00Z" in timestamps


# =============================================================================
# Task 4: CSV Escaping Tests (AC: #4)
# =============================================================================


class TestCSVEscaping:
    """Tests for CSV escaping and encoding."""

    def test_values_with_commas_quoted(
        self, edge_case_report_data: ReportData
    ) -> None:
        """Test values containing commas are properly quoted."""
        exporter = CSVExporter()
        result = exporter.export(edge_case_report_data)
        
        # CSV with commas in values should still parse correctly
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Find the row with commas in target
        targets = [row[2] for row in rows[1:]]
        assert "http://example.com/path,with,commas" in targets

    def test_values_with_quotes_escaped(
        self, edge_case_report_data: ReportData
    ) -> None:
        """Test values containing quotes are escaped (doubled)."""
        exporter = CSVExporter()
        result = exporter.export(edge_case_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Find the row with quotes in evidence
        descriptions = [row[3] for row in rows[1:]]
        assert any('"quotes"' in d for d in descriptions)

    def test_values_with_newlines_quoted(
        self, edge_case_report_data: ReportData
    ) -> None:
        """Test values containing newlines are properly quoted."""
        exporter = CSVExporter()
        result = exporter.export(edge_case_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Find the row with newlines
        descriptions = [row[3] for row in rows[1:]]
        assert any("Line 1\nLine 2" in d for d in descriptions)

    def test_unicode_characters_preserved(
        self, edge_case_report_data: ReportData
    ) -> None:
        """Test Unicode characters preserved (emojis, CJK, etc.)."""
        exporter = CSVExporter()
        result = exporter.export(edge_case_report_data)
        
        # Check emojis and CJK characters are preserved
        assert "🔥" in result
        assert "中文" in result
        assert "日本語" in result
        assert "한국어" in result

    def test_none_values_render_empty(
        self, edge_case_report_data: ReportData
    ) -> None:
        """Test None values render as empty string."""
        exporter = CSVExporter()
        result = exporter.export(edge_case_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Find the row with all None values (edge-002)
        # It should have empty strings for None fields
        none_row = None
        for row in rows[1:]:
            if all(cell == "" for cell in row[:4]):  # severity, type, target, evidence
                none_row = row
                break
        
        assert none_row is not None, "Should have a row with empty values for None fields"


# =============================================================================
# Task 5: ExcelExporter Class Tests (AC: #1, #2, #5)
# =============================================================================


class TestExcelExporterInit:
    """Tests for ExcelExporter initialization."""

    def test_init_creates_instance(self) -> None:
        """Test ExcelExporter.__init__() initializes correctly."""
        exporter = ExcelExporter()
        assert exporter is not None
        assert isinstance(exporter, ExcelExporter)


class TestExcelExporterExport:
    """Tests for ExcelExporter.export() method."""

    def test_export_returns_bytes(
        self, sample_report_data: ReportData
    ) -> None:
        """Test export(report_data) returns Excel bytes."""
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_to_file(
        self, sample_report_data: ReportData, tmp_path: Path
    ) -> None:
        """Test export(report_data, output_path) writes to file."""
        exporter = ExcelExporter()
        output_path = tmp_path / "findings.xlsx"
        
        result = exporter.export(sample_report_data, output_path=output_path)
        
        assert output_path.exists()
        assert isinstance(result, bytes)
        content = output_path.read_bytes()
        assert content == result

    def test_export_can_be_opened_by_openpyxl(
        self, sample_report_data: ReportData
    ) -> None:
        """Test Excel file can be opened by openpyxl."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        # Load workbook from bytes
        workbook = load_workbook(io.BytesIO(result))
        assert workbook is not None
        assert len(workbook.sheetnames) >= 1


# =============================================================================
# Task 6: Excel Formatting Tests (AC: #5)
# =============================================================================


class TestExcelFormatting:
    """Tests for Excel formatting (headers, auto-filter)."""

    def test_header_row_is_bold(
        self, sample_report_data: ReportData
    ) -> None:
        """Test Excel has formatted header row (bold)."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Check header cells are bold
        for cell in ws[1]:
            assert cell.font.bold is True, f"Header cell {cell.coordinate} should be bold"

    def test_auto_filter_enabled(
        self, sample_report_data: ReportData
    ) -> None:
        """Test header row has auto-filter enabled."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Check auto-filter is set
        assert ws.auto_filter.ref is not None

    def test_column_widths_reasonable(
        self, sample_report_data: ReportData
    ) -> None:
        """Test column widths are reasonable (not default narrow)."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Check column A (severity) has width >= 10
        assert ws.column_dimensions["A"].width >= 10

    def test_worksheet_named_findings(
        self, sample_report_data: ReportData
    ) -> None:
        """Test worksheet is named 'Findings'."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        
        assert "Findings" in workbook.sheetnames

    def test_all_data_rows_present(
        self, sample_report_data: ReportData
    ) -> None:
        """Test all data rows present under header."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Count rows: header + 5 findings = 6 rows
        row_count = ws.max_row
        assert row_count == 6


# =============================================================================
# Task 7: Excel Column Mapping Tests (AC: #3)
# =============================================================================


class TestExcelColumnMapping:
    """Tests for Excel column mapping."""

    def test_columns_match_csv(
        self, sample_report_data: ReportData
    ) -> None:
        """Test columns match CSV: severity, type, target, description, timestamp."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        headers = [cell.value for cell in ws[1]]
        assert headers == ["severity", "type", "target", "description", "timestamp"]

    def test_each_finding_produces_one_row(
        self, sample_report_data: ReportData
    ) -> None:
        """Test each finding produces one row (starting row 2)."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Row 2 should have data from first finding
        assert ws.cell(row=2, column=1).value == "critical"
        assert ws.cell(row=2, column=2).value == "sqli"

    def test_severity_values_preserved(
        self, sample_report_data: ReportData
    ) -> None:
        """Test severity values preserved."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        severities = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "critical" in severities
        assert "high" in severities

    def test_timestamp_values_as_strings(
        self, sample_report_data: ReportData
    ) -> None:
        """Test timestamp values preserved as strings (not Excel dates)."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Check timestamp is string format
        timestamp = ws.cell(row=2, column=5).value
        assert isinstance(timestamp, str)
        assert "2026-02-12" in timestamp


# =============================================================================
# Task 8: Edge Case Tests (AC: #6)
# =============================================================================


class TestEdgeCases:
    """Tests for edge cases."""

    def test_empty_findings_csv_header_only(
        self, empty_report_data: ReportData
    ) -> None:
        """Test empty findings produces valid CSV output with header only."""
        exporter = CSVExporter()
        result = exporter.export(empty_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Should have header only
        assert len(rows) == 1
        assert rows[0] == ["severity", "type", "target", "description", "timestamp"]

    def test_empty_findings_excel_header_only(
        self, empty_report_data: ReportData
    ) -> None:
        """Test empty findings produces valid Excel output with header only."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(empty_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Should have header only
        assert ws.max_row == 1

    def test_none_fields_handled_gracefully_csv(
        self, edge_case_report_data: ReportData
    ) -> None:
        """Test findings with None/missing fields handled gracefully in CSV."""
        exporter = CSVExporter()
        result = exporter.export(edge_case_report_data)
        
        # Should not raise exception
        assert isinstance(result, str)

    def test_none_fields_handled_gracefully_excel(
        self, edge_case_report_data: ReportData
    ) -> None:
        """Test findings with None/missing fields handled gracefully in Excel."""
        exporter = ExcelExporter()
        result = exporter.export(edge_case_report_data)
        
        # Should not raise exception
        assert isinstance(result, bytes)

    def test_datetime_objects_handled_csv(
        self, datetime_report_data: ReportData
    ) -> None:
        """Test findings with datetime objects (not strings) in CSV."""
        exporter = CSVExporter()
        result = exporter.export(datetime_report_data)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Timestamp should be converted to ISO string
        timestamp = rows[1][4]
        assert "2026-02-12" in timestamp

    def test_datetime_objects_handled_excel(
        self, datetime_report_data: ReportData
    ) -> None:
        """Test findings with datetime objects (not strings) in Excel."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(datetime_report_data)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Timestamp should be string format
        timestamp = ws.cell(row=2, column=5).value
        assert "2026-02-12" in str(timestamp)

    def test_long_evidence_text_not_truncated(
        self, sample_report_data: ReportData
    ) -> None:
        """Test findings with extremely long evidence text."""
        # Create finding with very long evidence
        long_evidence = "A" * 10000
        findings = (
            {
                "severity": "high",
                "type": "test",
                "target": "http://example.com",
                "evidence": long_evidence,
                "timestamp": "2026-02-12T06:00:00Z",
            },
        )
        report_data = ReportData(
            engagement_id="eng-long-001",
            title="Long Text Test",
            start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
            end_time=None,
            scope={"targets": [], "exclusions": []},
            findings=findings,
            timeline_events=(),
        )
        
        exporter = CSVExporter()
        result = exporter.export(report_data)
        
        # Evidence should not be truncated
        assert long_evidence in result

    def test_performance_1000_findings(self) -> None:
        """Test 1000+ findings (performance)."""
        # Create 1000 findings
        findings = tuple(
            {
                "id": f"finding-{i}",
                "severity": "medium",
                "type": "test",
                "target": f"http://example.com/{i}",
                "evidence": f"Finding {i} evidence",
                "timestamp": "2026-02-12T06:00:00Z",
            }
            for i in range(1000)
        )
        report_data = ReportData(
            engagement_id="eng-perf-001",
            title="Performance Test",
            start_time=datetime(2026, 2, 12, 6, 0, 0, tzinfo=timezone.utc),
            end_time=None,
            scope={"targets": [], "exclusions": []},
            findings=findings,
            timeline_events=(),
        )
        
        # CSV export
        csv_exporter = CSVExporter()
        csv_result = csv_exporter.export(report_data)
        
        reader = csv.reader(io.StringIO(csv_result))
        rows = list(reader)
        assert len(rows) == 1001  # Header + 1000 findings
        
        # Excel export
        excel_exporter = ExcelExporter()
        excel_result = excel_exporter.export(report_data)
        assert len(excel_result) > 0


# =============================================================================
# Task 9: Extended Columns Tests (AC: #3)
# =============================================================================


class TestExtendedColumns:
    """Tests for optional extended columns."""

    def test_extended_columns_csv(
        self, sample_report_data: ReportData
    ) -> None:
        """Test optional extended columns: agent_id, tool, topic, attck_id."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data, extended=True)
        
        reader = csv.reader(io.StringIO(result))
        header = next(reader)
        
        expected = [
            "severity", "type", "target", "description", "timestamp",
            "agent_id", "tool", "topic", "attck_id"
        ]
        assert header == expected

    def test_extended_true_includes_columns_csv(
        self, sample_report_data: ReportData
    ) -> None:
        """Test export(report_data, extended=True) includes extended columns."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data, extended=True)
        
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        
        # Check extended values exist
        assert len(rows[1]) == 9  # 5 standard + 4 extended

    def test_extended_false_uses_minimal_csv(
        self, sample_report_data: ReportData
    ) -> None:
        """Test export(report_data, extended=False) uses minimal columns."""
        exporter = CSVExporter()
        result = exporter.export(sample_report_data, extended=False)
        
        reader = csv.reader(io.StringIO(result))
        header = next(reader)
        
        assert len(header) == 5

    def test_extended_columns_excel(
        self, sample_report_data: ReportData
    ) -> None:
        """Test extended columns in Excel."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data, extended=True)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        headers = [cell.value for cell in ws[1]]
        expected = [
            "severity", "type", "target", "description", "timestamp",
            "agent_id", "tool", "topic", "attck_id"
        ]
        assert headers == expected

    def test_extended_true_includes_columns_excel(
        self, sample_report_data: ReportData
    ) -> None:
        """Test export(report_data, extended=True) includes extended columns in Excel."""
        from openpyxl import load_workbook
        
        exporter = ExcelExporter()
        result = exporter.export(sample_report_data, extended=True)
        
        workbook = load_workbook(io.BytesIO(result))
        ws = workbook.active
        
        # Check agent_id value in column F (6)
        agent_id = ws.cell(row=2, column=6).value
        assert agent_id == "recon-agent-1"


# =============================================================================
# Task 10: Convenience Functions Tests
# =============================================================================


class TestConvenienceFunctions:
    """Tests for convenience export functions."""

    def test_export_findings_csv(
        self, sample_report_data: ReportData
    ) -> None:
        """Test export_findings_csv function."""
        result = export_findings_csv(sample_report_data)
        assert isinstance(result, str)
        assert "severity,type,target,description,timestamp" in result

    def test_export_findings_csv_to_file(
        self, sample_report_data: ReportData, tmp_path: Path
    ) -> None:
        """Test export_findings_csv with output_path."""
        output_path = tmp_path / "findings.csv"
        result = export_findings_csv(sample_report_data, output_path=output_path)
        
        assert output_path.exists()
        assert isinstance(result, str)

    def test_export_findings_xlsx(
        self, sample_report_data: ReportData
    ) -> None:
        """Test export_findings_xlsx function."""
        result = export_findings_xlsx(sample_report_data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_findings_xlsx_to_file(
        self, sample_report_data: ReportData, tmp_path: Path
    ) -> None:
        """Test export_findings_xlsx with output_path."""
        output_path = tmp_path / "findings.xlsx"
        result = export_findings_xlsx(sample_report_data, output_path=output_path)
        
        assert output_path.exists()
        assert isinstance(result, bytes)


# =============================================================================
# Task: Validation and Error Handling Tests
# =============================================================================


class TestValidationAndErrorHandling:
    """Tests for input validation and error handling."""

    def test_csv_export_invalid_type_raises_type_error(self) -> None:
        """Test CSVExporter.export raises TypeError for invalid input."""
        exporter = CSVExporter()
        
        with pytest.raises(TypeError, match="must have a 'findings' attribute"):
            exporter.export("not a report data")  # type: ignore

    def test_excel_export_invalid_type_raises_type_error(self) -> None:
        """Test ExcelExporter.export raises TypeError for invalid input."""
        exporter = ExcelExporter()
        
        with pytest.raises(TypeError, match="must have a 'findings' attribute"):
            exporter.export("not a report data")  # type: ignore

    def test_csv_export_findings_string_raises_value_error(self) -> None:
        """Test CSVExporter.export raises ValueError when findings is a string."""
        class FakeReportData:
            findings = "not a list"
        
        exporter = CSVExporter()
        
        with pytest.raises(ValueError, match="must be an iterable of dicts"):
            exporter.export(FakeReportData())  # type: ignore

    def test_excel_export_findings_string_raises_value_error(self) -> None:
        """Test ExcelExporter.export raises ValueError when findings is a string."""
        class FakeReportData:
            findings = "not a list"
        
        exporter = ExcelExporter()
        
        with pytest.raises(ValueError, match="must be an iterable of dicts"):
            exporter.export(FakeReportData())  # type: ignore

    def test_csv_export_invalid_path_raises_ioerror(
        self, sample_report_data: ReportData
    ) -> None:
        """Test CSVExporter.export raises IOError for invalid path."""
        exporter = CSVExporter()
        
        with pytest.raises(IOError, match="parent directory does not exist"):
            exporter.export(sample_report_data, output_path=Path("/nonexistent/dir/file.csv"))

    def test_excel_export_invalid_path_raises_ioerror(
        self, sample_report_data: ReportData
    ) -> None:
        """Test ExcelExporter.export raises IOError for invalid path."""
        exporter = ExcelExporter()
        
        with pytest.raises(IOError, match="parent directory does not exist"):
            exporter.export(sample_report_data, output_path=Path("/nonexistent/dir/file.xlsx"))

    def test_write_to_file_permission_error(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test _write_to_file raises IOError for permission denied."""
        from cyberred.storage.csv_excel_exporter import _write_to_file
        
        test_path = tmp_path / "test.txt"
        
        # Mock Path.write_text to raise PermissionError
        def mock_write_text(*args: Any, **kwargs: Any) -> None:
            raise PermissionError("Permission denied")
        
        monkeypatch.setattr(Path, "write_text", mock_write_text)
        
        with pytest.raises(IOError, match="permission denied"):
            _write_to_file(test_path, "test content", encoding="utf-8")

    def test_write_to_file_generic_oserror(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test _write_to_file raises IOError for generic OSError."""
        from cyberred.storage.csv_excel_exporter import _write_to_file
        
        test_path = tmp_path / "test.txt"
        
        # Mock Path.write_text to raise generic OSError
        def mock_write_text(*args: Any, **kwargs: Any) -> None:
            raise OSError("Disk full or other OS error")
        
        monkeypatch.setattr(Path, "write_text", mock_write_text)
        
        with pytest.raises(IOError, match="Disk full or other OS error"):
            _write_to_file(test_path, "test content", encoding="utf-8")

    def test_write_to_file_bytes_permission_error(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test _write_to_file raises IOError for permission denied with bytes."""
        from cyberred.storage.csv_excel_exporter import _write_to_file
        
        test_path = tmp_path / "test.bin"
        
        # Mock Path.write_bytes to raise PermissionError
        def mock_write_bytes(*args: Any, **kwargs: Any) -> None:
            raise PermissionError("Permission denied")
        
        monkeypatch.setattr(Path, "write_bytes", mock_write_bytes)
        
        with pytest.raises(IOError, match="permission denied"):
            _write_to_file(test_path, b"test content")
